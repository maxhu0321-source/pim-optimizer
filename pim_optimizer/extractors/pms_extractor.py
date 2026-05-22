"""PMS 标准代码表 (.xlsx) 数据提取器"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from ..config import PMS_14_COLS, PMS_14_DATA_START
from ..models import PMSData, RoomType


def _clean_code(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        value = int(value)
    return str(value).strip().upper()


def _to_int(value: Any) -> int:
    if value is None or value == "":
        return 0
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def extract_pms(file_path: Path) -> PMSData:
    """从 PMS 标准代码表提取房型数据"""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        data = PMSData()

        # 1.4 酒店客房
        if "1.4 酒店客房" in wb.sheetnames:
            ws = wb["1.4 酒店客房"]
            for row_num, row in enumerate(ws.iter_rows(min_row=PMS_14_DATA_START, values_only=True), start=PMS_14_DATA_START):
                if not row or len(row) < 7:
                    continue

                # 检测空行或分隔行（房号列表区域），停止读取
                seq = row[PMS_14_COLS["seq"]]
                code = _clean_code(row[PMS_14_COLS["code"]])
                if seq is None and not code:
                    # 连续空行 = 房型汇总区结束
                    break

                if not code:
                    continue

                # 排除虚拟房（不提供给希尔顿渠道）
                hilton = _clean_code(row[PMS_14_COLS["hilton_code"]])
                if "不提供" in str(row[PMS_14_COLS["hilton_code"]] or ""):
                    continue

                count = _to_int(row[PMS_14_COLS["count"]])
                if count <= 0:
                    continue

                rt = RoomType(
                    source="PMS-1.4",
                    code=code,
                    name_cn=str(row[PMS_14_COLS["name_cn"]] or "").strip(),
                    name_en=str(row[PMS_14_COLS["name_en"]] or "").strip(),
                    count=count,
                    hilton_code=hilton,
                    row=row_num,
                )
                data.room_types.append(rt)

        return data
    finally:
        wb.close()
